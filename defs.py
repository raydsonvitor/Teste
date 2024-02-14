import pandas as pd
from tkinter import *
from datetime import datetime, date
import openpyxl as op
from CTkMessagebox import CTkMessagebox
from defsofdefs import *

arquivo_name = 'defs.py'
despesas = ['CONSUMÍVEIS', 'ALUGUEL', 'ÁGUA', 'LUZ', 'INTERNET','FUNDOS', 'ESTORNO', 'OUTRO']
form_pgmt_saida = ['DINHEIRO', 'BANCO', 'CARTÃO']
form_pgmt_entrada = ['DINHEIRO', 'PIX', 'DÉBITO', 'CRÉDITO','MENSAL']
valor_limite_saida = 9999
valor_limite_entrada = 9999

def Lin():
    print('-------------------------------')

def Obter_faturamento(periodo):
    #abrindo a database
    book = op.load_workbook(r'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_23.xlsx')
    sheet = book[periodo]
    rows = list(sheet.values)
    counter = soma = 0
    for row in rows:
        counter+=1
        if counter >=2:
            if row[5] != None:
                soma+=float(row[5])
    return soma

def Obter_faturamento_por_barbeiro_diario(data, periodo):
    try:
        ano = '20'+periodo[3:]
        #abrindo a database
        book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        #identificando as rows certas
        profissionais = ObterListaProfissionais()
        profissional1 = profissional2 = profissional3 = profissional4 = 0
        for row in rows:
            #filtrando data
            if row[1] == data:
                #filtrando o profissional
                if row[2] == profissionais[0]:
                    profissional1 += float(row[5])
                elif row[2] == profissionais[1]:
                    profissional2 += float(row[5])
                elif row[2] == profissionais[2]:
                    profissional3 += float(row[5])
                elif row[2] == profissionais[3]:
                    profissional4 += float(row[5])
        return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
    except:
        print('Erro ao calcular o faturamento diário por barbeiro!')
        return ['0,00', '0,00', '0,00', '0,00']

def Obter_faturamento_por_barbeiro_semanal(data, periodo):
    try:
        ano = '20'+periodo[3:]
        ##obter a lista de datas pertencentes à semana
        #obter o indice do weekday de hoje
        weekday_indice = date.today().weekday()
        #subtrair do dia da data o indice do weekday
        dia = data.split('-')[0]
        mes = data.split('-')[1]
        print(f'data:{dia}-{mes}. weekday indice: {weekday_indice}')
        if (int(dia) - weekday_indice) > 0:#Para o caso de os dias da semana estarem contidos somente no mes atual
            print('A subtração foi maior')
            last_monday_day = int(dia) - weekday_indice
            #formar a lista de datas desde a ultima segunda feira
            lista_datas = []
            if weekday_indice == 0:
                print('O indice foi igual a 0')
                lista_datas.append(data)
                book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
                sheet = book[periodo]
                rows = list(sheet.values)
                #identificando as rows certas
                profissionais = ObterListaProfissionais()
                profissional1 = profissional2 = profissional3 = profissional4 = 0
                for row in rows:
                    #filtrando data
                    if row[1] in lista_datas:
                        #filtrando o profissional
                        if row[2] == profissionais[0]:
                            profissional1 += float(row[5])
                        elif row[2] == profissionais[1]:
                            profissional2 += float(row[5])
                        elif row[2] == profissionais[2]:
                            profissional3 += float(row[5])
                        elif row[2] == profissionais[3]:
                            profissional4 += float(row[5])
                return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']

            else:
                print('O indice foi maior do que 0')
                last_day = last_monday_day
                for vez in range(weekday_indice + 1):
                    variable_data = str(last_day)+'-'+mes
                    lista_datas.append(variable_data)
                    last_day += 1
            book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
            sheet = book[periodo]
            rows = list(sheet.values)
            #identificando as rows certas
            profissionais = ObterListaProfissionais()
            profissional1 = profissional2 = profissional3 = profissional4 = 0
            for row in rows:
                #filtrando data
                if row[1] in lista_datas:
                    #filtrando o profissional
                    if row[2] == profissionais[0]:
                        profissional1 += float(row[5])
                    elif row[2] == profissionais[1]:
                        profissional2 += float(row[5])
                    elif row[2] == profissionais[2]:
                        profissional3 += float(row[5])
                    elif row[2] == profissionais[3]:
                        profissional4 += float(row[5])
            return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']

        elif (int(dia) - weekday_indice) <= 0:#Para o caso de haverem dias da semana no mes anterior
            print('A subtração foi menor')
            mes = str(int(periodo.split('-')[0])-1)
            ano_abrev = periodo.split('-')[1]
            print(mes, ano)
            dias_mes = DiasMes(mes, ano_abrev)
            print(f'dias mes {dias_mes}')
            last_monday_day = dias_mes + int(dia) - weekday_indice
            print(f'last monday day = dias mes + dia - weekdindice')
            print(f'{last_monday_day} = {dias_mes} + {dia} - {weekday_indice}')
            #formar a lista de datas desde a ultima segunda feira
            lista_datas = []
            if weekday_indice == 0:
                print('O indice foi igual a 0')
                lista_datas.append(data)
                return lista_datas
            else:
                print('O indice foi maior do que 0')
                last_day = last_monday_day
                for vez in range(weekday_indice + 1):
                    if last_day > dias_mes:
                        last_day = 1
                        mes = int(mes)+1
                    variable_data = str(last_day)+'-'+str(mes)
                    lista_datas.append(variable_data)
                    last_day += 1
            print(lista_datas)
            book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
            print(periodo)
            mes = str(int(periodo.split('-')[0]))
            ano = periodo.split('-')[1]
            sheet = book[f'{Zero_adder(str(int(mes)-1))}-{ano_abrev}']
            print('aqui começa o periodo'+f'{str(int(mes)-1)}-{ano_abrev}')
            rows = list(sheet.values)
            #identificando as rows certas
            profissionais = ObterListaProfissionais()
            profissional1 = profissional2 = profissional3 = profissional4 = 0
            for row in rows:#periodo anterior
                #filtrando data
                if row[1] in lista_datas:
                    #filtrando o profissional
                    if row[2] == profissionais[0]:
                        profissional1 += float(row[5])
                    elif row[2] == profissionais[1]:
                        profissional2 += float(row[5])
                    elif row[2] == profissionais[2]:
                        profissional3 += float(row[5])
                    elif row[2] == profissionais[3]:
                        profissional4 += float(row[5])
            sheet = book[f'{Zero_adder(mes)}-{ano}']
            rows = list(sheet.values)
            for row in rows:#periodo atual
                #filtrando data
                if row[1] in lista_datas:
                    #filtrando o profissional
                    if row[2] == profissionais[0]:
                        profissional1 += float(row[5])
                    elif row[2] == profissionais[1]:
                        profissional2 += float(row[5])
                    elif row[2] == profissionais[2]:
                        profissional3 += float(row[5])
                    elif row[2] == profissionais[3]:
                        profissional4 += float(row[5])
            print([f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}'])
            return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
    except:
        print(f'Erro durante a execução da função Obter_faturamento_por_barbeiro_semanal no arquivo {arquivo_name}')
        return ['0.00', '0.00', '0.00', '0.00']

def Obter_faturamento_por_barbeiro_mensal(periodo):
    try:
        ano = '20'+periodo[3:]
        #abrindo a database
        book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        #identificando as rows certas
        profissionais = ObterListaProfissionais()
        profissional1 = profissional2 = profissional3 = profissional4 = 0
        for row in rows:
            #filtrando o profissional
            if row[2] == profissionais[0]:
                profissional1 += float(row[5])
            elif row[2] == profissionais[1]:
                profissional2 += float(row[5])
            elif row[2] == profissionais[2]:
                profissional3 += float(row[5])
            elif row[2] == profissionais[3]:
                profissional4 += float(row[5])
        return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
    except:
        print('Erro ao calcular o faturamento mensal por barbeiro!')
        return ['0,00', '0,00', '0,00', '0,00']

def Obter_total_entrada_dinheiro(data, periodo):
    try:
        ano = '20'+periodo[3:]
        #abrindo o databse
        book = op.load_workbook(fr'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        total = 0
        for row in rows:
            if row[1] == data:
                if row[4] == 'DINHEIRO' and row[6]==None:
                    total += float(row[5])
        return total
    except:
        print('Erro durante a execução da função Obter_total_dinheiro (defs.py).')
        return float('00.00')

def Obter_total_saida_dinheiro(data, periodo):
    try:
        ano = '20'+periodo[3:] 
        #abrindo o databse
        book = op.load_workbook(fr'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        total = 0
        for row in rows:
            if row[1] == data:
                if row[4] == 'DINHEIRO' and row[5]==None:
                    total += float(row[6])
        print(total)
        return total
    except:
        print('Erro durante a execução da função Obter_total_dinheiro (defs.py).')
        return float('00.00')

def Obter_ultimo_caixa():
    try:
        arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\caixa.txt', 'r')
        saldo = float(arquivo.readline())
        return saldo
    except:
        print('Erro durante a execução da função Obter_ultimo_caixa (defs.py).')
        return float(f'00.00')

def Obter_caixa(data, periodo):
    try:
        arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\caixa.txt', 'r')
        saldo = float(arquivo.readline())
        ultimo_caixa = Obter_ultimo_caixa()
        total_entrada_dinheiro = Obter_total_entrada_dinheiro(data, periodo)
        total_saida_dinheiro = Obter_total_saida_dinheiro(data, periodo)
        caixa = ultimo_caixa + total_entrada_dinheiro - total_saida_dinheiro
        return caixa
    except:
        print('Ocorreu um erro durante a execução da função Obter_caixa (defs.py)')
        return float(f'00.00')

def GetFaturamentoDia():
    pass

def GetPeriodo(colunas_database):
#try:
    #definições importantes
    ano_hoje = date.today().year
    mes_hoje = date.today().month

    #pegando o periodo anual do txt
    arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\periodo_anual.txt', 'r')
    periodo_anual = arquivo.readline()
    
    #comparando os periodos anuais do txt e do computador
    if ano_hoje > int(periodo_anual):
        #criando um novo arquivo referente ao novo periodo anual
        mes_hoje = Zero_adder(mes_hoje)
        novoperiodo = f'{(mes_hoje)}-{str(ano_hoje)[2:]}'
        wb = op.Workbook()
        sheet = wb.active
        sheet.title = novoperiodo
        sheet.append(colunas_database)
        wb.save(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano_hoje}.xlsx')
        print(f'Novo arquivo excell criado referente ao novo periodo anual {ano_hoje}')
        #atualizando os txt's dos periodos
        arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\periodo_anual.txt', 'w')
        arquivo.write(str(ano_hoje))
        print(f'Novo periodo anual setado: {ano_hoje}')
        arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\periodo_mensal.txt', 'w')
        arquivo.write(str(mes_hoje))
        print(f'Novo periodo mensal setado: {mes_hoje}')
        print(f'Periodo setado: {novoperiodo}')
        return novoperiodo
    else:
        print(f'O periodo anual permance o mesmo: {periodo_anual}')

    #pegando o periodo mensal no txt
    arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\periodo_mensal.txt', 'r')
    periodo_mensal = arquivo.readline()

    #comparando os periodos mensais do txt e do computador
    if mes_hoje - int(periodo_mensal) != 0:
        mes_hoje = Zero_adder(mes_hoje)
        novoperiodo = f'{(mes_hoje)}-{periodo_anual[2:]}'
        #criando uma nova sheet referente ao novo periodo mensal
        book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{periodo_anual}.xlsx')#abrindo o db
        book.create_sheet(novoperiodo)#criando nova sheet correspondente ao novo periodo mensal
        #acrescentando os titulo das colunas do database
        sheet = book[novoperiodo]
        sheet.append(colunas_database)
        book.save(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{periodo_anual}.xlsx')
        print(f'Uma nova sheet com o nome {novoperiodo} foi criada no arquivo nw_barbearia_{periodo_anual}.xlsx')
        #atualizar no txt
        arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\periodo_mensal.txt', 'w')
        arquivo.write(str(mes_hoje))
        print(f'Novo periodo mensal setado: {mes_hoje}')
        print(f'Periodo setado: {novoperiodo}')
        return novoperiodo
    else:
        print(f'O periodo mensal permance o mesmo: {periodo_mensal}')
    #retornando o periodo que nada mudou
    return f'{periodo_mensal}-{periodo_anual[2:]}'
#except:
    print('Ocorreu um erro na função GetPeriodo em defs.py')

def GetData():
    dia = str(date.today().day)
    mes = str(date.today().month)
    data = f'{dia}-{mes}'
    return data

def GetHora():
    hr = str(datetime.now())[11:16]
    return hr

def GetLastId(path, periodo, lista):
    book = op.load_workbook(path)
    sheet = book[periodo]
    rows = list(sheet.values)
    if rows[-1] == lista:#p\ o caso de nao haver movimentações
        return 0
    return rows[-1][0]

def Check_0(profissional, servico, form_pgmt, valor):#checkar inputs da area de registro de entradas
    profissionais = ObterListaProfissionais()
    if profissional == 'PROFISSO.' or profissional not in profissionais:
        print('Opção escolhida ''profissional'' não foi aceita')
        msg = 'Verifique a opção Profissional e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg)
        return False
    if servico == '':
        print('Opção escolhida ''serviço'' não foi aceita')
        msg = 'Selecione um serviço válido e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg)
        return False
    if form_pgmt == 'FORM. PGMT.' or form_pgmt not in form_pgmt_entrada:
        print('Opção escolhida ''forma de pagamento'' não foi aceita')
        msg = 'Verifique a opção Forma de Pagamento e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg)
        return False
    if valor == '' or float(valor) > valor_limite_entrada:
        print('Opção escolhida ''Valor'' não foi aceita')
        msg = 'Verifique a opção Valor e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg)
        return False
    return True

def Check_1(despesa, form_pgmt, valor):#checkar inuts da area de registro de saidas
    #row = [id, data, '', despesa, form_pgmt, '', valor]
    if despesa == 'DESPESA' or despesa not in despesas:
        print('Opção escolhida ''despesa'' não foi aceita')
        msg = 'Verifique a opção Despesa e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Saída', message=msg)
        return False
    if form_pgmt == 'FORM. PGMT.' or form_pgmt not in form_pgmt_saida:
        print('Opção escolhida ''form pgmt'' não foi aceita')
        msg = 'Verifique a opção Form. Pgmt. e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Saída', message=msg)
        return False
    if valor == '' or float(valor) > valor_limite_saida:
        print('Opção escolhida ''valor'' não foi aceita')
        msg = 'Verifique a opção Valor. e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Saída', message=msg)
        return False
    return True
        
def ObterListaProfissionais():
    try:
        arquivo = open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\profissionais.txt', 'r')
        linha = (arquivo.readline())
        profissionais = linha.split(';')
        return profissionais
    except:
        print('Erro ao capturar a lista de profissionais!')
        return ['###', '###', '###', '###']

def Soma(lista):
    try:
        soma = 0
        for item in lista:
            soma+=float(item)
        return soma
    except:
        print('Houve um erro na função Soma')

def CloseApp(master):
    print('fechando app...')
    master.destroy()
    print('App fechado!')

def DiasMes(mes, ano):
    try:   
        month = int(mes)
        if month == 2:#se é o mês fevereiro
            ano = int('20'+ano)
            if ano % 4 == 0:#se for bissexto
                return 29
            else:
                return 28
        if month % 2 != 0:#se o mes for ímpar
            if month <= 7:
                return 31
        if month % 2 == 0:#se o mes for par
            if month >= 8:
                return 31
        else:
            return 30

    except:
        print(f'Erro durante a execução da função DiasMes no arquivo {arquivo_name}')

def Get_faturamento_dia_by_formpgmt(ano, periodo, data):
    try:
        #abrindo a database
        book = op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        dinheiro = cartao = pix = 0
        #identificando as rows certas
        for row in rows:
            #filtrando data
            if row[1] == data:
                #filtrando o por form pgmt
                if row[4] == 'dinheiro'.upper():
                    dinheiro += float(row[5])
                if row[4] == 'débito'.upper() or row[4]=='crédito'.upper():
                    cartao += float(row[5])
                if row[4] == 'pix'.upper():
                    pix += float(row[5])
        return [dinheiro, cartao, pix]
    except:
        print('Erro na função Get_faturamento_dia')
        return ['0.00', '0.00','0.00']